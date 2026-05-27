import paho.mqtt.client as mqtt
import time
import json
import re
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)
from mqtt_init import *

# Student: Yakira Siboni, ID: 208499426
DHT_topic = 'pr/home/5976397/sts'
samples = []
TARGET_SAMPLES = 20

def on_log(client, userdata, level, buf):
    print("log: " + buf)

def on_connect(client, userdata, flags, rc):
    if rc == 0:
        print("Connected OK")
        client.subscribe(DHT_topic, qos=0)
        print(f"Subscribed to: {DHT_topic}")
        print(f"Collecting {TARGET_SAMPLES} DHT samples (this may take ~10 minutes)...\n")
    else:
        print("Bad connection, code =", rc)

def on_disconnect(client, userdata, flags, rc=0):
    print("Disconnected, code =", str(rc))

def on_message(client, userdata, msg):
    global samples
    m_decode = str(msg.payload.decode("utf-8", "ignore"))
    print(f"[{len(samples)+1}/{TARGET_SAMPLES}] Raw: {m_decode}")

    # Parse "Temperature: XX.X Humidity: YY.Y"
    temp_match = re.search(r'Temperature:\s*([\d.]+)', m_decode)
    hum_match  = re.search(r'Humidity:\s*([\d.]+)', m_decode)

    if temp_match and hum_match:
        temp = float(temp_match.group(1))
        hum  = float(hum_match.group(1))
        timestamp = time.strftime('%H:%M:%S')
        samples.append({'time': timestamp, 'temperature': temp, 'humidity': hum})
        print(f"    -> Temp: {temp}°C  Hum: {hum}%")

        if len(samples) >= TARGET_SAMPLES:
            print(f"\nCollected {TARGET_SAMPLES} samples. Saving to Excel...")
            save_to_excel()
            client.disconnect()
    else:
        print("    -> Could not parse DHT data, skipping.")

def save_to_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import LineChart, Reference

        wb = Workbook()
        ws = wb.active
        ws.title = "DHT Data"

        # Header
        headers = ['#', 'Time', 'Temperature (°C)', 'Humidity (%)']
        header_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
        thin = Side(style="thin", color="AAAAAA")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        # Data rows
        for i, s in enumerate(samples, 1):
            row = [i, s['time'], s['temperature'], s['humidity']]
            alt_fill = PatternFill("solid", start_color="DCE6F1" if i % 2 == 0 else "FFFFFF", end_color="DCE6F1" if i % 2 == 0 else "FFFFFF")
            for col, val in enumerate(row, 1):
                cell = ws.cell(row=i+1, column=col, value=val)
                cell.font = Font(name="Arial", size=11)
                cell.alignment = center
                cell.border = border
                cell.fill = alt_fill

        # Column widths
        for col, width in zip(['A','B','C','D'], [6, 12, 20, 16]):
            ws.column_dimensions[col].width = width

        # Line chart - Temperature
        chart_temp = LineChart()
        chart_temp.title = "Temperature over Time"
        chart_temp.style = 10
        chart_temp.y_axis.title = "Temperature (°C)"
        chart_temp.x_axis.title = "Sample"
        data_temp = Reference(ws, min_col=3, min_row=1, max_row=len(samples)+1)
        chart_temp.add_data(data_temp, titles_from_data=True)
        chart_temp.shape = 4
        ws.add_chart(chart_temp, "F2")

        # Line chart - Humidity
        chart_hum = LineChart()
        chart_hum.title = "Humidity over Time"
        chart_hum.style = 10
        chart_hum.y_axis.title = "Humidity (%)"
        chart_hum.x_axis.title = "Sample"
        data_hum = Reference(ws, min_col=4, min_row=1, max_row=len(samples)+1)
        chart_hum.add_data(data_hum, titles_from_data=True)
        chart_hum.shape = 4
        ws.add_chart(chart_hum, "F18")

        filename = "dht_data_yakira_siboni.xlsx"
        wb.save(filename)
        print(f"Saved: {filename}")
    except Exception as e:
        print(f"Excel save failed: {e}")
        # Fallback: print to console
        print("\n--- DHT Data ---")
        for s in samples:
            print(f"{s['time']}  Temp: {s['temperature']}°C  Hum: {s['humidity']}%")

import random
r = random.randrange(1, 10000000)
client = mqtt.Client(mqtt.CallbackAPIVersion.VERSION1, f"IOT_dht_collector_{r}", clean_session=True)
client.on_connect = on_connect
client.on_disconnect = on_disconnect
client.on_message = on_message
client.username_pw_set(username, password)

print(f"Connecting to broker {broker_ip}:{broker_port}")
client.connect(broker_ip, int(broker_port))
client.loop_start()

# Wait up to 15 minutes
timeout = 15 * 60
start = time.time()
while len(samples) < TARGET_SAMPLES and (time.time() - start) < timeout:
    time.sleep(1)

client.loop_stop()
client.disconnect()
print("Test 1 complete.")
